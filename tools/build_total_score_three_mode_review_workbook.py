from __future__ import annotations

import argparse
import csv
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from mk8_local_play.project_paths import PROJECT_ROOT


VARIANTS = (
    ("baseline", "Baseline"),
    ("transition_only", "Transition-only"),
    ("stable_hint", "Transition + stable-hint"),
)

TOURNAMENT_KEY = ["Video", "Race", "Track", "Position"]
FINAL_KEY = ["VideoName", "Position", "PlayerName"]
TRACE_KEY = ["RaceClass", "Race"]
TOURNAMENT_FIELDS = [
    "Player",
    "Character",
    "Character Roster",
    "Race Points",
    "Total Before Race",
    "Total After Race",
    "Position After Race",
    "Needs Review",
    "Review Reason",
    "Counts Toward Totals",
    "Scoring Note",
]
FINAL_FIELDS = [
    "Races",
    "TotalPoints",
    "Character",
    "CharacterRosterName",
    "Points Segment 1",
    "Points Segment 2",
    "Points Segment 3",
]
TRACE_FIELDS = [
    "Candidate Frame",
    "Score Hit Frame",
    "Race Anchor Frame",
    "Actual Race Anchor Frame",
    "Transition Frame",
    "Points Anchor Frame",
    "Actual Points Anchor Frame",
    "Stable Total Frame",
    "Total Anchor Frame",
    "Actual Total Anchor Frame",
]
OCR_STAGE_FIELDS = [
    "PlayerName",
    "FixPlayerName",
    "Character",
    "DetectedTotalScore",
    "OldTotalScore",
    "NewTotalScore",
    "NameConfidence",
    "ReviewNeeded",
    "ReviewReason",
    "IdentityResolutionMethod",
]

YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
RED_FILL = PatternFill("solid", fgColor="F4CCCC")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")


def parse_args():
    parser = argparse.ArgumentParser(description="Build one human-review workbook for baseline, transition-only, and stable-hint runs.")
    parser.add_argument("--root", required=True, help="Root created by run_total_score_three_mode_review.py")
    parser.add_argument(
        "--output",
        default=str(PROJECT_ROOT / ".codex_tmp" / "top30_three_mode_review.xlsx"),
        help="Workbook output path",
    )
    return parser.parse_args()


def _load_csv(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, keep_default_na=False, sep=None, engine="python")
    df.columns = [str(column).lstrip("\ufeff") for column in df.columns]
    return df


def _latest_matching_file(folder: Path, suffix: str) -> Path:
    matches = sorted(folder.glob(f"*_{suffix}"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not matches:
        raise FileNotFoundError(f"No *_{suffix} file found in {folder}")
    return matches[0]


def _load_manifest(manifest_path: Path) -> dict[str, str]:
    result: dict[str, str] = {}
    with manifest_path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.reader(handle):
            if len(row) >= 2 and row[0] and row[1]:
                result[str(row[0])] = str(row[1])
    return result


def _normalize(value) -> str:
    if pd.isna(value):
        return ""
    return str(value)


def _coerce_int(value):
    text = _normalize(value).strip()
    if not text:
        return ""
    try:
        return int(float(text))
    except ValueError:
        return text


def _load_trace_map(mode_dir: Path) -> dict[tuple[str, int], dict]:
    trace_path = mode_dir / "total_score_frame_trace.csv"
    if not trace_path.exists():
        return {}
    df = _load_csv(trace_path)
    trace_map = {}
    for row in df.to_dict(orient="records"):
        trace_map[(str(row.get("VideoLabel") or row.get("RaceClass") or ""), int(row.get("Race") or 0))] = row
    return trace_map


def _load_ocr_stage_map(mode_dir: Path, stage_name: str) -> dict[tuple[str, int, int], dict]:
    root = mode_dir / "OCR_Tracing" / "identity_stages"
    stage_map: dict[tuple[str, int, int], dict] = {}
    if not root.exists():
        return stage_map
    for video_dir in root.iterdir():
        if not video_dir.is_dir():
            continue
        stage_path = video_dir / f"{stage_name}.csv"
        if not stage_path.exists():
            continue
        df = _load_csv(stage_path)
        for row in df.to_dict(orient="records"):
            stage_map[(str(row.get("RaceClass") or video_dir.name), int(row.get("RaceIDNumber") or 0), int(row.get("RacePosition") or 0))] = row
    return stage_map


def _find_bundle_paths(mode_dir: Path, video_label: str, race: int) -> dict[str, str]:
    race_dir = mode_dir / "Frames" / video_label / f"Race_{int(race):03d}"
    race_bundle = race_dir / "2RaceScore"
    total_bundle = race_dir / "3TotalScore"
    race_anchor = sorted(race_bundle.glob("anchor_*")) if race_bundle.exists() else []
    total_anchor = sorted(total_bundle.glob("anchor_*")) if total_bundle.exists() else []
    return {
        "Race Bundle Dir": str(race_bundle) if race_bundle.exists() else "",
        "Race Anchor Image": str(race_anchor[0]) if race_anchor else "",
        "Total Bundle Dir": str(total_bundle) if total_bundle.exists() else "",
        "Total Anchor Image": str(total_anchor[0]) if total_anchor else "",
    }


def _compare_fields(base_row: dict, compare_row: dict, fields: list[str]) -> tuple[str, int]:
    changed = []
    for field in fields:
        if _normalize(base_row.get(field, "")) != _normalize(compare_row.get(field, "")):
            changed.append(field)
    return ", ".join(changed), len(changed)


def _write_sheet_headers(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def _autosize(ws, max_width: int = 48) -> None:
    for column_cells in ws.columns:
        length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        ws.column_dimensions[column_letter].width = min(max(10, length + 2), max_width)


def _add_false_formatting(ws) -> None:
    red_fill = RED_FILL
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        if str(cell.value).startswith("EQ "):
            col_letter = get_column_letter(col_idx)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                FormulaRule(formula=[f'{col_letter}2=FALSE'], fill=red_fill),
            )


def _apply_file_hyperlink(cell, *, label: str) -> bool:
    value = str(cell.value or "").strip()
    if not value:
        return False
    try:
        target = Path(value).resolve()
    except Exception:
        return False
    if not target.exists():
        return False
    cell.hyperlink = target.as_uri()
    cell.value = label
    cell.style = "Hyperlink"
    return True


def build_workbook(root: Path, output_path: Path) -> Path:
    manifests = {}
    tournament = {}
    finals = {}
    traces = {}
    raw_stages = {}
    final_stages = {}
    for mode_key, _mode_label in VARIANTS:
        mode_dir = root / mode_key
        manifests[mode_key] = _load_manifest(mode_dir / "manifest.csv")
        tournament[mode_key] = _load_csv(_latest_matching_file(mode_dir, "Tournament_Results.csv"))
        finals[mode_key] = _load_csv(_latest_matching_file(mode_dir, "Final_Standings.csv"))
        traces[mode_key] = _load_trace_map(mode_dir)
        raw_stages[mode_key] = _load_ocr_stage_map(mode_dir, "raw_ocr_input")
        final_stages[mode_key] = _load_ocr_stage_map(mode_dir, "final_identity_export_input")

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_headers = [
        "Mode",
        "Description",
        "Total",
        "Extract",
        "OCR",
        "Tournament CSV",
        "Final CSV",
        "Tournament Changed Rows vs Baseline",
        "Tournament Changed Fields vs Baseline",
        "Final Changed Rows vs Baseline",
        "Final Changed Fields vs Baseline",
        "Videos With Tournament Diffs",
        "Videos With Final Diffs",
        "What To Evaluate",
        "Hyperlinks Verified",
    ]
    _write_sheet_headers(summary_ws, summary_headers)

    base_t = tournament["baseline"]
    base_f = finals["baseline"]
    for mode_key, mode_label in VARIANTS:
        mode_t = tournament[mode_key]
        mode_f = finals[mode_key]
        tournament_merged = base_t.merge(mode_t, on=TOURNAMENT_KEY, how="outer", suffixes=("_BASE", "_MODE"))
        tournament_compare_cols = sorted({col[:-5] for col in tournament_merged.columns if col.endswith("_BASE")})
        tournament_diffs = []
        for row in tournament_merged.to_dict(orient="records"):
            changed_fields = [
                col for col in tournament_compare_cols
                if _normalize(row.get(f"{col}_BASE", "")) != _normalize(row.get(f"{col}_MODE", ""))
            ]
            if changed_fields:
                tournament_diffs.append((str(row.get("Video", "")), changed_fields))

        final_merged = base_f.merge(mode_f, on=FINAL_KEY, how="outer", suffixes=("_BASE", "_MODE"))
        final_compare_cols = sorted({col[:-5] for col in final_merged.columns if col.endswith("_BASE")})
        final_diffs = []
        for row in final_merged.to_dict(orient="records"):
            changed_fields = [
                col for col in final_compare_cols
                if _normalize(row.get(f"{col}_BASE", "")) != _normalize(row.get(f"{col}_MODE", ""))
            ]
            if changed_fields:
                final_diffs.append((str(row.get("VideoName", "")), changed_fields))

        if mode_key == "baseline":
            evaluate_text = "Reference only"
        elif mode_key == "transition_only":
            evaluate_text = "Check low-res Amber->Ambor name drift and one review-flag removal"
        else:
            evaluate_text = "Broad drift review; faster but not yet safe by default"

        summary_ws.append(
            [
                mode_key,
                mode_label,
                manifests[mode_key].get("TotalProcessingTime", ""),
                manifests[mode_key].get("ExtractTime", ""),
                manifests[mode_key].get("OcrTime", ""),
                _latest_matching_file(root / mode_key, "Tournament_Results.csv").name,
                _latest_matching_file(root / mode_key, "Final_Standings.csv").name,
                0 if mode_key == "baseline" else len(tournament_diffs),
                "" if mode_key == "baseline" else ", ".join(sorted({field for _video, fields in tournament_diffs for field in fields})),
                0 if mode_key == "baseline" else len(final_diffs),
                "" if mode_key == "baseline" else ", ".join(sorted({field for _video, fields in final_diffs for field in fields})),
                "" if mode_key == "baseline" else ", ".join(sorted({video for video, _fields in tournament_diffs})),
                "" if mode_key == "baseline" else ", ".join(sorted({video for video, _fields in final_diffs})),
                evaluate_text,
                "",
            ]
        )

    tournament_ws = wb.create_sheet("Tournament All Rows")
    tournament_headers = [
        *TOURNAMENT_KEY,
        "Changed Transition?",
        "Changed Stable?",
        "Transition Changed Fields",
        "Stable Changed Fields",
        "Transition False Count",
        "Stable False Count",
    ]
    for mode_key, mode_label in VARIANTS:
        tournament_headers.extend(
            [
                f"{mode_label} Race Bundle Dir",
                f"{mode_label} Race Anchor",
                f"{mode_label} Total Bundle Dir",
                f"{mode_label} Total Anchor",
                *[f"{mode_label} {field}" for field in TRACE_FIELDS],
                *[f"{mode_label} RAW {field}" for field in OCR_STAGE_FIELDS],
                *[f"{mode_label} FINAL {field}" for field in OCR_STAGE_FIELDS],
                *[f"{mode_label} {field}" for field in TOURNAMENT_FIELDS],
            ]
        )
        if mode_key != "baseline":
            tournament_headers.extend([f"EQ {mode_label} {field}" for field in TOURNAMENT_FIELDS])
    _write_sheet_headers(tournament_ws, tournament_headers)

    base_rows = {
        tuple(row[key] for key in TOURNAMENT_KEY): row
        for row in base_t.to_dict(orient="records")
    }
    all_keys = sorted(
        {
            tuple(row[key] for key in TOURNAMENT_KEY)
            for mode_df in tournament.values()
            for row in mode_df.to_dict(orient="records")
        },
        key=lambda item: (str(item[0]), int(item[1]), str(item[2]), int(item[3])),
    )
    transition_rows = {
        tuple(row[key] for key in TOURNAMENT_KEY): row
        for row in tournament["transition_only"].to_dict(orient="records")
    }
    stable_rows = {
        tuple(row[key] for key in TOURNAMENT_KEY): row
        for row in tournament["stable_hint"].to_dict(orient="records")
    }

    for key in all_keys:
        base_row = base_rows.get(key, {})
        transition_row = transition_rows.get(key, {})
        stable_row = stable_rows.get(key, {})
        transition_fields, transition_false_count = _compare_fields(base_row, transition_row, TOURNAMENT_FIELDS)
        stable_fields, stable_false_count = _compare_fields(base_row, stable_row, TOURNAMENT_FIELDS)
        row_values = [
            *key,
            bool(transition_fields),
            bool(stable_fields),
            transition_fields,
            stable_fields,
            transition_false_count,
            stable_false_count,
        ]
        for mode_key, mode_label in VARIANTS:
            mode_row_map = {
                "baseline": base_row,
                "transition_only": transition_row,
                "stable_hint": stable_row,
            }
            mode_row = mode_row_map[mode_key]
            video = str(mode_row.get("Video", key[0]))
            race = int(mode_row.get("Race", key[1]) or 0)
            bundle_paths = _find_bundle_paths(root / mode_key, video, race)
            trace_row = traces[mode_key].get((video, race), {})
            raw_row = raw_stages[mode_key].get((video, race, int(mode_row.get("Position", key[3]) or key[3])), {})
            final_row = final_stages[mode_key].get((video, race, int(mode_row.get("Position", key[3]) or key[3])), {})
            row_values.extend(
                [
                    bundle_paths["Race Bundle Dir"],
                    bundle_paths["Race Anchor Image"],
                    bundle_paths["Total Bundle Dir"],
                    bundle_paths["Total Anchor Image"],
                    *[_coerce_int(trace_row.get(field, "")) for field in TRACE_FIELDS],
                    *[_normalize(raw_row.get(field, "")) for field in OCR_STAGE_FIELDS],
                    *[_normalize(final_row.get(field, "")) for field in OCR_STAGE_FIELDS],
                    *[_normalize(mode_row.get(field, "")) for field in TOURNAMENT_FIELDS],
                ]
            )
            if mode_key != "baseline":
                row_values.extend([None for _ in TOURNAMENT_FIELDS])
        tournament_ws.append(row_values)

    # Insert formulas and hyperlinks for tournament sheet.
    header_index = {str(tournament_ws.cell(row=1, column=col).value): col for col in range(1, tournament_ws.max_column + 1)}
    hyperlink_verified = 0
    for row_idx in range(2, tournament_ws.max_row + 1):
        for mode_label in ("Transition-only", "Transition + stable-hint"):
            for field in TOURNAMENT_FIELDS:
                base_col = header_index[f"Baseline {field}"]
                compare_col = header_index[f"{mode_label} {field}"]
                eq_col = header_index[f"EQ {mode_label} {field}"]
                base_letter = get_column_letter(base_col)
                compare_letter = get_column_letter(compare_col)
                tournament_ws.cell(row=row_idx, column=eq_col).value = f"={base_letter}{row_idx}={compare_letter}{row_idx}"
        for mode_label in ("Baseline", "Transition-only", "Transition + stable-hint"):
            for link_field in ("Race Bundle Dir", "Race Anchor", "Total Bundle Dir", "Total Anchor"):
                col_idx = header_index[f"{mode_label} {link_field}"]
                cell = tournament_ws.cell(row=row_idx, column=col_idx)
                if _apply_file_hyperlink(cell, label=link_field):
                    hyperlink_verified += 1

    final_ws = wb.create_sheet("Final All Rows")
    final_headers = [
        *FINAL_KEY,
        "Changed Transition?",
        "Changed Stable?",
        "Transition Changed Fields",
        "Stable Changed Fields",
        "Transition False Count",
        "Stable False Count",
    ]
    for mode_key, mode_label in VARIANTS:
        final_headers.extend([f"{mode_label} {field}" for field in FINAL_FIELDS])
        if mode_key != "baseline":
            final_headers.extend([f"EQ {mode_label} {field}" for field in FINAL_FIELDS])
    _write_sheet_headers(final_ws, final_headers)

    base_final_rows = {tuple(row[key] for key in FINAL_KEY): row for row in base_f.to_dict(orient="records")}
    trans_final_rows = {tuple(row[key] for key in FINAL_KEY): row for row in finals["transition_only"].to_dict(orient="records")}
    stable_final_rows = {tuple(row[key] for key in FINAL_KEY): row for row in finals["stable_hint"].to_dict(orient="records")}
    final_keys = sorted(
        {
            tuple(row[key] for key in FINAL_KEY)
            for mode_df in finals.values()
            for row in mode_df.to_dict(orient="records")
        },
        key=lambda item: (str(item[0]), int(item[1]), str(item[2])),
    )

    for key in final_keys:
        base_row = base_final_rows.get(key, {})
        trans_row = trans_final_rows.get(key, {})
        stable_row = stable_final_rows.get(key, {})
        trans_fields, trans_count = _compare_fields(base_row, trans_row, FINAL_FIELDS)
        stable_fields, stable_count = _compare_fields(base_row, stable_row, FINAL_FIELDS)
        row_values = [*key, bool(trans_fields), bool(stable_fields), trans_fields, stable_fields, trans_count, stable_count]
        for mode_key, _mode_label in VARIANTS:
            mode_row = {
                "baseline": base_row,
                "transition_only": trans_row,
                "stable_hint": stable_row,
            }[mode_key]
            row_values.extend([_normalize(mode_row.get(field, "")) for field in FINAL_FIELDS])
            if mode_key != "baseline":
                row_values.extend([None for _ in FINAL_FIELDS])
        final_ws.append(row_values)

    final_header_index = {str(final_ws.cell(row=1, column=col).value): col for col in range(1, final_ws.max_column + 1)}
    for row_idx in range(2, final_ws.max_row + 1):
        for mode_label in ("Transition-only", "Transition + stable-hint"):
            for field in FINAL_FIELDS:
                base_col = final_header_index[f"Baseline {field}"]
                compare_col = final_header_index[f"{mode_label} {field}"]
                eq_col = final_header_index[f"EQ {mode_label} {field}"]
                base_letter = get_column_letter(base_col)
                compare_letter = get_column_letter(compare_col)
                final_ws.cell(row=row_idx, column=eq_col).value = f"={base_letter}{row_idx}={compare_letter}{row_idx}"

    changed_ws = wb.create_sheet("Changed Only")
    changed_headers = [
        "Scope",
        "Video",
        "Race",
        "Track/Player",
        "Key",
        "Transition Changed Fields",
        "Stable Changed Fields",
        "Transition Summary",
        "Stable Summary",
    ]
    _write_sheet_headers(changed_ws, changed_headers)
    for row_idx in range(2, tournament_ws.max_row + 1):
        if tournament_ws.cell(row=row_idx, column=5).value or tournament_ws.cell(row=row_idx, column=6).value:
            changed_ws.append(
                [
                    "Tournament",
                    tournament_ws.cell(row=row_idx, column=1).value,
                    tournament_ws.cell(row=row_idx, column=2).value,
                    tournament_ws.cell(row=row_idx, column=3).value,
                    f"Pos {tournament_ws.cell(row=row_idx, column=4).value}",
                    tournament_ws.cell(row=row_idx, column=7).value,
                    tournament_ws.cell(row=row_idx, column=8).value,
                    f"{tournament_ws.cell(row=row_idx, column=5).value} | {tournament_ws.cell(row=row_idx, column=9).value}",
                    f"{tournament_ws.cell(row=row_idx, column=6).value} | {tournament_ws.cell(row=row_idx, column=10).value}",
                ]
            )
    for row_idx in range(2, final_ws.max_row + 1):
        if final_ws.cell(row=row_idx, column=4).value or final_ws.cell(row=row_idx, column=5).value:
            changed_ws.append(
                [
                    "Final",
                    final_ws.cell(row=row_idx, column=1).value,
                    "",
                    final_ws.cell(row=row_idx, column=3).value,
                    f"Pos {final_ws.cell(row=row_idx, column=2).value}",
                    final_ws.cell(row=row_idx, column=6).value,
                    final_ws.cell(row=row_idx, column=7).value,
                    f"{final_ws.cell(row=row_idx, column=4).value} | {final_ws.cell(row=row_idx, column=8).value}",
                    f"{final_ws.cell(row=row_idx, column=5).value} | {final_ws.cell(row=row_idx, column=9).value}",
                ]
            )

    for ws in (summary_ws, tournament_ws, final_ws, changed_ws):
        _autosize(ws)
    _add_false_formatting(tournament_ws)
    _add_false_formatting(final_ws)
    for row_idx in range(2, summary_ws.max_row + 1):
        summary_ws.cell(row=row_idx, column=summary_headers.index("Hyperlinks Verified") + 1).value = hyperlink_verified

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main():
    args = parse_args()
    output_path = build_workbook(Path(args.root), Path(args.output))
    print(output_path)


if __name__ == "__main__":
    main()
