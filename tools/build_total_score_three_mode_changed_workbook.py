from __future__ import annotations

import argparse
import csv
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from mk8_local_play.project_paths import PROJECT_ROOT


VARIANTS = (
    ("baseline", "Baseline"),
    ("transition_only", "Transition-only"),
    ("stable_hint", "Transition + stable-hint"),
)

TOURNAMENT_KEY = ["Video", "Race", "Track", "Position"]
FINAL_KEY = ["VideoName", "Position", "PlayerName"]
TRACE_FIELDS = [
    "Candidate Frame",
    "Score Hit Frame",
    "Race Anchor Frame",
    "Actual Race Anchor Frame",
    "Transition Frame",
    "Points Anchor Frame",
    "Actual Points Anchor Frame",
    "Stable Total Frame",
    "Total Anchor Frame",
    "Actual Total Anchor Frame",
]
TOURNAMENT_FIELDS = [
    "Player",
    "Character",
    "Character Roster",
    "Race Points",
    "Total Before Race",
    "Total After Race",
    "Position After Race",
    "Needs Review",
    "Review Reason",
]
FINAL_FIELDS = [
    "Races",
    "TotalPoints",
    "Character",
    "CharacterRosterName",
    "Points Segment 1",
    "Points Segment 2",
    "Points Segment 3",
]
OCR_FIELDS = [
    "PlayerName",
    "FixPlayerName",
    "Character",
    "DetectedTotalScore",
    "OldTotalScore",
    "NewTotalScore",
    "NameConfidence",
    "ReviewNeeded",
    "ReviewReason",
    "IdentityResolutionMethod",
]

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
RED_FILL = PatternFill("solid", fgColor="F4CCCC")


def parse_args():
    parser = argparse.ArgumentParser(description="Build a compact changed-only workbook for the three-mode Total Score review.")
    parser.add_argument("--root", required=True, help="Root created by run_total_score_three_mode_review.py")
    parser.add_argument(
        "--output",
        default=str(PROJECT_ROOT / ".codex_tmp" / "top30_three_mode_review" / "top30_three_mode_changed_cases.xlsx"),
        help="Workbook output path",
    )
    return parser.parse_args()


def _load_csv(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, keep_default_na=False, sep=None, engine="python")
    df.columns = [str(column).lstrip("\ufeff") for column in df.columns]
    return df


def _latest_matching_file(folder: Path, suffix: str) -> Path:
    matches = sorted(folder.glob(f"*_{suffix}"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not matches:
        raise FileNotFoundError(f"No *_{suffix} file found in {folder}")
    return matches[0]


def _normalize(value) -> str:
    if pd.isna(value):
        return ""
    return str(value)


def _coerce_int(value):
    text = _normalize(value).strip()
    if not text:
        return ""
    try:
        return int(float(text))
    except ValueError:
        return text


def _load_trace_map(mode_dir: Path) -> dict[tuple[str, int], dict]:
    trace_path = mode_dir / "total_score_frame_trace.csv"
    if not trace_path.exists():
        return {}
    df = _load_csv(trace_path)
    return {
        (str(row.get("VideoLabel") or row.get("RaceClass") or ""), int(row.get("Race") or 0)): row
        for row in df.to_dict(orient="records")
    }


def _load_ocr_stage_map(mode_dir: Path, stage_name: str) -> dict[tuple[str, int, int], dict]:
    root = mode_dir / "OCR_Tracing" / "identity_stages"
    result: dict[tuple[str, int, int], dict] = {}
    if not root.exists():
        return result
    for video_dir in root.iterdir():
        if not video_dir.is_dir():
            continue
        stage_path = video_dir / f"{stage_name}.csv"
        if not stage_path.exists():
            continue
        df = _load_csv(stage_path)
        for row in df.to_dict(orient="records"):
            result[(str(row.get("RaceClass") or video_dir.name), int(row.get("RaceIDNumber") or 0), int(row.get("RacePosition") or 0))] = row
    return result


def _find_bundle_paths(mode_dir: Path, video_label: str, race: int) -> dict[str, str]:
    race_dir = mode_dir / "Frames" / video_label / f"Race_{int(race):03d}"
    race_bundle = race_dir / "2RaceScore"
    total_bundle = race_dir / "3TotalScore"
    race_anchor = sorted(race_bundle.glob("anchor_*")) if race_bundle.exists() else []
    total_anchor = sorted(total_bundle.glob("anchor_*")) if total_bundle.exists() else []
    return {
        "Race Bundle Dir": str(race_bundle) if race_bundle.exists() else "",
        "Race Anchor Image": str(race_anchor[0]) if race_anchor else "",
        "Total Bundle Dir": str(total_bundle) if total_bundle.exists() else "",
        "Total Anchor Image": str(total_anchor[0]) if total_anchor else "",
    }


def _write_headers(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def _autosize(ws, max_width: int = 48) -> None:
    for column_cells in ws.columns:
        max_len = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column_letter].width = min(max(10, max_len + 2), max_width)


def _add_false_formatting(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        if str(ws.cell(row=1, column=col_idx).value).startswith("EQ "):
            col_letter = get_column_letter(col_idx)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                FormulaRule(formula=[f"{col_letter}2=FALSE"], fill=RED_FILL),
            )


def _apply_file_hyperlink(cell, *, label: str) -> bool:
    value = str(cell.value or "").strip()
    if not value:
        return False
    try:
        target = Path(value).resolve()
    except Exception:
        return False
    if not target.exists():
        return False
    cell.hyperlink = target.as_uri()
    cell.value = label
    cell.style = "Hyperlink"
    return True


def build_changed_workbook(root: Path, output_path: Path) -> Path:
    tournament = {}
    finals = {}
    traces = {}
    raw_stages = {}
    final_stages = {}
    for mode_key, _mode_label in VARIANTS:
        mode_dir = root / mode_key
        tournament[mode_key] = _load_csv(_latest_matching_file(mode_dir, "Tournament_Results.csv"))
        finals[mode_key] = _load_csv(_latest_matching_file(mode_dir, "Final_Standings.csv"))
        traces[mode_key] = _load_trace_map(mode_dir)
        raw_stages[mode_key] = _load_ocr_stage_map(mode_dir, "raw_ocr_input")
        final_stages[mode_key] = _load_ocr_stage_map(mode_dir, "final_identity_export_input")

    base_t = tournament["baseline"]
    transition_t = tournament["transition_only"]
    stable_t = tournament["stable_hint"]
    base_f = finals["baseline"]
    transition_f = finals["transition_only"]
    stable_f = finals["stable_hint"]

    base_t_map = {tuple(row[key] for key in TOURNAMENT_KEY): row for row in base_t.to_dict(orient="records")}
    transition_t_map = {tuple(row[key] for key in TOURNAMENT_KEY): row for row in transition_t.to_dict(orient="records")}
    stable_t_map = {tuple(row[key] for key in TOURNAMENT_KEY): row for row in stable_t.to_dict(orient="records")}
    base_f_map = {tuple(row[key] for key in FINAL_KEY): row for row in base_f.to_dict(orient="records")}
    transition_f_map = {tuple(row[key] for key in FINAL_KEY): row for row in transition_f.to_dict(orient="records")}
    stable_f_map = {tuple(row[key] for key in FINAL_KEY): row for row in stable_f.to_dict(orient="records")}

    changed_t_keys = []
    for key in sorted(set(base_t_map) | set(transition_t_map) | set(stable_t_map), key=lambda item: (str(item[0]), int(item[1]), str(item[2]), int(item[3]))):
        base_row = base_t_map.get(key, {})
        trans_row = transition_t_map.get(key, {})
        stable_row = stable_t_map.get(key, {})
        if any(_normalize(base_row.get(field, "")) != _normalize(trans_row.get(field, "")) for field in TOURNAMENT_FIELDS) or any(
            _normalize(base_row.get(field, "")) != _normalize(stable_row.get(field, "")) for field in TOURNAMENT_FIELDS
        ):
            changed_t_keys.append(key)

    changed_f_keys = []
    for key in sorted(set(base_f_map) | set(transition_f_map) | set(stable_f_map), key=lambda item: (str(item[0]), int(item[1]), str(item[2]))):
        base_row = base_f_map.get(key, {})
        trans_row = transition_f_map.get(key, {})
        stable_row = stable_f_map.get(key, {})
        if any(_normalize(base_row.get(field, "")) != _normalize(trans_row.get(field, "")) for field in FINAL_FIELDS) or any(
            _normalize(base_row.get(field, "")) != _normalize(stable_row.get(field, "")) for field in FINAL_FIELDS
        ):
            changed_f_keys.append(key)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Cases"
    _write_headers(
        summary_ws,
        [
            "Scope",
            "Video",
            "Race",
            "Track/Player",
            "Position",
            "Transition Changed Fields",
            "Stable Changed Fields",
            "Human Focus",
        ],
    )

    tournament_ws = wb.create_sheet("Tournament Changed")
    t_headers = [
        *TOURNAMENT_KEY,
        "Transition Changed Fields",
        "Stable Changed Fields",
    ]
    for mode_key, mode_label in VARIANTS:
        t_headers.extend(
            [
                f"{mode_label} Race Bundle Dir",
                f"{mode_label} Race Anchor",
                f"{mode_label} Total Bundle Dir",
                f"{mode_label} Total Anchor",
                *[f"{mode_label} {field}" for field in TRACE_FIELDS],
                *[f"{mode_label} RAW {field}" for field in OCR_FIELDS],
                *[f"{mode_label} FINAL {field}" for field in OCR_FIELDS],
                *[f"{mode_label} {field}" for field in TOURNAMENT_FIELDS],
            ]
        )
        if mode_key != "baseline":
            t_headers.extend([f"EQ {mode_label} {field}" for field in TOURNAMENT_FIELDS])
    _write_headers(tournament_ws, t_headers)

    for key in changed_t_keys:
        base_row = base_t_map.get(key, {})
        trans_row = transition_t_map.get(key, {})
        stable_row = stable_t_map.get(key, {})
        trans_changed = ", ".join(field for field in TOURNAMENT_FIELDS if _normalize(base_row.get(field, "")) != _normalize(trans_row.get(field, "")))
        stable_changed = ", ".join(field for field in TOURNAMENT_FIELDS if _normalize(base_row.get(field, "")) != _normalize(stable_row.get(field, "")))
        row_values = [*key, trans_changed, stable_changed]
        for mode_key, mode_label in VARIANTS:
            mode_row = {
                "baseline": base_row,
                "transition_only": trans_row,
                "stable_hint": stable_row,
            }[mode_key]
            video = str(mode_row.get("Video", key[0]))
            race = int(mode_row.get("Race", key[1]) or 0)
            bundle = _find_bundle_paths(root / mode_key, video, race)
            trace = traces[mode_key].get((video, race), {})
            raw = raw_stages[mode_key].get((video, race, int(mode_row.get("Position", key[3]) or key[3])), {})
            final = final_stages[mode_key].get((video, race, int(mode_row.get("Position", key[3]) or key[3])), {})
            row_values.extend(
                [
                    bundle["Race Bundle Dir"],
                    bundle["Race Anchor Image"],
                    bundle["Total Bundle Dir"],
                    bundle["Total Anchor Image"],
                    *[_coerce_int(trace.get(field, "")) for field in TRACE_FIELDS],
                    *[_normalize(raw.get(field, "")) for field in OCR_FIELDS],
                    *[_normalize(final.get(field, "")) for field in OCR_FIELDS],
                    *[_normalize(mode_row.get(field, "")) for field in TOURNAMENT_FIELDS],
                ]
            )
            if mode_key != "baseline":
                row_values.extend([None for _ in TOURNAMENT_FIELDS])
        tournament_ws.append(row_values)
        focus = "Review low-res OCR / identity drift" if "Player" in (trans_changed + stable_changed) else "Review flag / reasoning change"
        summary_ws.append(["Tournament", key[0], key[1], key[2], key[3], trans_changed, stable_changed, focus])

    t_header_index = {str(tournament_ws.cell(row=1, column=col).value): col for col in range(1, tournament_ws.max_column + 1)}
    for row_idx in range(2, tournament_ws.max_row + 1):
        for mode_label in ("Transition-only", "Transition + stable-hint"):
            for field in TOURNAMENT_FIELDS:
                base_col = t_header_index[f"Baseline {field}"]
                compare_col = t_header_index[f"{mode_label} {field}"]
                eq_col = t_header_index[f"EQ {mode_label} {field}"]
                tournament_ws.cell(row=row_idx, column=eq_col).value = (
                    f"={get_column_letter(base_col)}{row_idx}={get_column_letter(compare_col)}{row_idx}"
                )
        for mode_label in ("Baseline", "Transition-only", "Transition + stable-hint"):
            for link_field in ("Race Bundle Dir", "Race Anchor", "Total Bundle Dir", "Total Anchor"):
                col_idx = t_header_index[f"{mode_label} {link_field}"]
                cell = tournament_ws.cell(row=row_idx, column=col_idx)
                _apply_file_hyperlink(cell, label=link_field)

    final_ws = wb.create_sheet("Final Changed")
    f_headers = [*FINAL_KEY, "Transition Changed Fields", "Stable Changed Fields"]
    for mode_key, mode_label in VARIANTS:
        f_headers.extend([f"{mode_label} {field}" for field in FINAL_FIELDS])
        if mode_key != "baseline":
            f_headers.extend([f"EQ {mode_label} {field}" for field in FINAL_FIELDS])
    _write_headers(final_ws, f_headers)

    for key in changed_f_keys:
        base_row = base_f_map.get(key, {})
        trans_row = transition_f_map.get(key, {})
        stable_row = stable_f_map.get(key, {})
        trans_changed = ", ".join(field for field in FINAL_FIELDS if _normalize(base_row.get(field, "")) != _normalize(trans_row.get(field, "")))
        stable_changed = ", ".join(field for field in FINAL_FIELDS if _normalize(base_row.get(field, "")) != _normalize(stable_row.get(field, "")))
        row_values = [*key, trans_changed, stable_changed]
        for mode_key, _mode_label in VARIANTS:
            mode_row = {
                "baseline": base_row,
                "transition_only": trans_row,
                "stable_hint": stable_row,
            }[mode_key]
            row_values.extend([_normalize(mode_row.get(field, "")) for field in FINAL_FIELDS])
            if mode_key != "baseline":
                row_values.extend([None for _ in FINAL_FIELDS])
        final_ws.append(row_values)
        summary_ws.append(["Final", key[0], "", key[2], key[1], trans_changed, stable_changed, "Review final totals / identity output"])

    f_header_index = {str(final_ws.cell(row=1, column=col).value): col for col in range(1, final_ws.max_column + 1)}
    for row_idx in range(2, final_ws.max_row + 1):
        for mode_label in ("Transition-only", "Transition + stable-hint"):
            for field in FINAL_FIELDS:
                base_col = f_header_index[f"Baseline {field}"]
                compare_col = f_header_index[f"{mode_label} {field}"]
                eq_col = f_header_index[f"EQ {mode_label} {field}"]
                final_ws.cell(row=row_idx, column=eq_col).value = (
                    f"={get_column_letter(base_col)}{row_idx}={get_column_letter(compare_col)}{row_idx}"
                )

    for ws in (summary_ws, tournament_ws, final_ws):
        _autosize(ws)
    _add_false_formatting(tournament_ws)
    _add_false_formatting(final_ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main():
    args = parse_args()
    output_path = build_changed_workbook(Path(args.root), Path(args.output))
    print(output_path)


if __name__ == "__main__":
    main()
