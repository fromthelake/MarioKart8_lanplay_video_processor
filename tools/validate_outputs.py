import argparse
import csv
import hashlib
import sys
from pathlib import Path

from openpyxl import load_workbook


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def compare_file_sets(base_dir: Path, current_dir: Path, prefix: str = ""):
    base_files = sorted(
        [path for path in base_dir.glob("*.png") if path.name.startswith(prefix)],
        key=lambda item: item.name,
    )
    current_files = sorted(
        [path for path in current_dir.glob("*.png") if path.name.startswith(prefix)],
        key=lambda item: item.name,
    )

    base_map = {path.name: sha256_file(path) for path in base_files}
    current_map = {path.name: sha256_file(path) for path in current_files}

    missing = sorted(set(base_map) - set(current_map))
    unexpected = sorted(set(current_map) - set(base_map))
    mismatched = sorted(
        name for name in (set(base_map) & set(current_map))
        if base_map[name] != current_map[name]
    )
    return missing, unexpected, mismatched


def load_csv_rows(csv_path: Path, race_class: str | None):
    with csv_path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.reader(handle))
    if not race_class:
        return rows
    if not rows:
        return rows
    header = rows[0]
    filtered = [header]
    filtered.extend(row for row in rows[1:] if row and row[0] == race_class)
    return filtered


def load_workbook_rows(xlsx_path: Path, race_class: str | None):
    workbook = load_workbook(xlsx_path, data_only=True)
    worksheet = workbook.active
    rows = []
    for row in worksheet.iter_rows(values_only=True):
        rows.append(["" if value is None else str(value) for value in row])
    if not race_class:
        return rows
    if not rows:
        return rows
    header = rows[0]
    filtered = [header]
    filtered.extend(row for row in rows[1:] if row and row[0] == race_class)
    return filtered


def align_rows_for_comparison(baseline_rows, current_rows):
    if not baseline_rows or not current_rows:
        return baseline_rows, current_rows
    baseline_header = baseline_rows[0]
    current_header = current_rows[0]
    shared_columns = [column for column in baseline_header if column in current_header]
    if shared_columns == baseline_header and len(current_header) == len(baseline_header):
        return baseline_rows, current_rows

    baseline_indices = [baseline_header.index(column) for column in shared_columns]
    current_indices = [current_header.index(column) for column in shared_columns]

    def _project(rows, indices):
        projected = [shared_columns]
        for row in rows[1:]:
            projected.append([row[index] if index < len(row) else "" for index in indices])
        return projected

    return _project(baseline_rows, baseline_indices), _project(current_rows, current_indices)


def print_list(label: str, values):
    print(f"{label}: {len(values)}")
    for value in values[:10]:
        print(f"  {value}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate current outputs against a stored baseline")
    parser.add_argument("--baseline-dir", required=True, help="Baseline directory containing Frames/ and Debug/")
    parser.add_argument("--current-dir", default="Output_Results", help="Current output directory")
    parser.add_argument("--prefix", default="", help="Filename prefix filter, for example Divisie_1")
    parser.add_argument("--race-class", default="", help="RaceClass filter for workbook comparison")
    parser.add_argument("--skip-score-frames", action="store_true", help="Skip annotated score-frame comparison")
    args = parser.parse_args()

    baseline_dir = Path(args.baseline_dir)
    current_dir = Path(args.current_dir)
    prefix = args.prefix
    race_class = args.race_class or None

    frames_missing, frames_unexpected, frames_mismatched = compare_file_sets(
        baseline_dir / "Frames",
        current_dir / "Frames",
        prefix=prefix,
    )
    print_list("frames_missing", frames_missing)
    print_list("frames_unexpected", frames_unexpected)
    print_list("frames_mismatched", frames_mismatched)

    score_missing = []
    score_unexpected = []
    score_mismatched = []
    if not args.skip_score_frames:
        score_prefix = f"annotated_{prefix}" if prefix else ""
        score_missing, score_unexpected, score_mismatched = compare_file_sets(
            baseline_dir / "Debug" / "Score_Frames",
            current_dir / "Debug" / "Score_Frames",
            prefix=score_prefix,
        )
        print_list("score_frames_missing", score_missing)
        print_list("score_frames_unexpected", score_unexpected)
        print_list("score_frames_mismatched", score_mismatched)

    baseline_rows = load_csv_rows(baseline_dir / "Tournament_Results.csv", race_class)
    current_rows = load_workbook_rows(current_dir / "Tournament_Results.xlsx", race_class)
    baseline_rows, current_rows = align_rows_for_comparison(baseline_rows, current_rows)
    workbook_match = baseline_rows == current_rows
    print(f"workbook_match: {workbook_match}")
    print(f"baseline_rows: {len(baseline_rows)}")
    print(f"current_rows: {len(current_rows)}")
    if not workbook_match:
        max_len = max(len(baseline_rows), len(current_rows))
        for index in range(max_len):
            base_row = baseline_rows[index] if index < len(baseline_rows) else None
            current_row = current_rows[index] if index < len(current_rows) else None
            if base_row != current_row:
                print(f"first_workbook_diff_index: {index}")
                print(f"baseline_row: {base_row}")
                print(f"current_row: {current_row}")
                break

    success = (
        not frames_missing
        and not frames_unexpected
        and not frames_mismatched
        and not score_missing
        and not score_unexpected
        and not score_mismatched
        and workbook_match
    )
    print(f"validation_passed: {success}")
    return 0 if success else 1


if __name__ == "__main__":
    sys.exit(main())
